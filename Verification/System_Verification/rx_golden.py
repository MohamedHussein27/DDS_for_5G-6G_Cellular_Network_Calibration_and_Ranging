"""
    Sponsor: Analog Devices, Inc. (ADI)
    Institution: Faculty of Engineering, Ain Shams University
    Project: DDS for 5G/6G Cellular Network Calibration and Ranging

    Module: rx_golden_model.py

    Description:
        Master wrapper for the full RX datapath golden model.
        Mirrors the structure of tx_golden_model.py exactly.

        Pipeline (mirrors RX_top.v):
            Stage 1 : FFT (4096-pt, bit-true, Radix-2 DIF)
            Stage 2 : Mid-path bit-reversal (ADDR_W=12, N=4096)
                        → output[k] = fft_out[bit_rev_12(k)]
            Stage 3 : ISAC Demux
                        → OFDM  : mid_rev_out[0..2047]   (communication signal)
                        → Radar : mid_rev_out[2048..4095] (range-profile path)
            Stage 4 : Conjugate multiply
                        → ref_conj_im = -ref_ram_im
                        → multiply(radar[k], re2=ref_ram_re[k], im2=ref_conj_im[k])
                        → RTL: re_out = sub_re[29:14] + sub_re[13]  (round)
            Stage 5 : Arithmetic right-shift >>>3 before IFFT
            Stage 6 : IFFT (2048-pt, bit-true, Radix-2 DIF)
            Stage 7 : Final bit-reversal (ADDR_W=11, N=2048)
                        → output[k] = ifft_out[bit_rev_11(k)]

        Fixed-point parameters:
            WL = 16, FL = 8 for FFT/demux/multiply stages
            FL = 5 for IFFT input  (after >>>3 shift)

        Outputs two pairs of integer arrays ready for UVM scoreboard:
            ofdm_out_re [2048]  — communication signal (real)
            ofdm_out_im [2048]  — communication signal (imag)
            radar_out_re[2048]  — range profile (real)
            radar_out_im[2048]  — range profile (imag)

    DEBUG DUMP:
        When debug_xlsx is not None, run_rx_pipeline() writes a workbook
        with one sheet per pipeline stage so you can compare against RTL
        at each boundary:

            Sheet "0_Inputs"         — rx_in and ref_ram samples
            Sheet "1_FFT_Out"        — 4096 FFT output bins
            Sheet "2_MidBitRev_Out"  — 4096 post-bit-reversal bins
            Sheet "3a_OFDM_Bins"     — 2048 OFDM demux output
            Sheet "3b_Radar_Bins"    — 2048 Radar demux output
            Sheet "4_ConjMul_Out"    — 2048 multiply output (before shift)
            Sheet "5_Shift_Out"      — 2048 after >>>3 shift (IFFT input)
            Sheet "6_IFFT_Out"       — 2048 IFFT output
            Sheet "7_FinalBitRev"    — 2048 final radar output
"""

import numpy as np
from fft_fixed  import radix2_dif_fft_fixed
from ifft_fixed import radix2_dif_ifft_fixed
from openpyxl                        import Workbook
from openpyxl.styles                 import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils                  import get_column_letter

# ===========================================================================
# PARAMETERS
# ===========================================================================
WL      = 16
FL_FFT  = 8
FL_IFFT = 5      # after >>>3 shift on the multiply output (Q8.8 >>> 3 → Q11.5)
N_FFT   = 4096
N_HALF  = 2048


# ===========================================================================
# HELPER FUNCTIONS
# ===========================================================================

def _sign16(v: int) -> int:
    """Sign-extend a raw 16-bit integer."""
    v = int(v) & 0xFFFF
    return v - 0x10000 if (v & 0x8000) else v


def _asr(val: int, shift: int) -> int:
    """Arithmetic shift right — mirrors Verilog >>> on signed wire."""
    return int(val) >> shift


def _bit_rev(k: int, width: int) -> int:
    """Reverse 'width' LSBs of k."""
    result = 0
    for _ in range(width):
        result = (result << 1) | (k & 1)
        k >>= 1
    return result


def _apply_bit_reversal(samples: np.ndarray, addr_w: int) -> np.ndarray:
    """
    Exact model of bit_reversal_pingpong:
        output[k] = input[bit_rev(k, addr_w)]
    """
    N   = len(samples)
    out = np.zeros(N, dtype=np.complex128)
    for k in range(N):
        out[k] = samples[_bit_rev(k, addr_w)]
    return out


def _multiply_rtl(re1: int, im1: int, ref_re: int, ref_im: int):
    """
    Exact integer replica of multiplier.v + RX_top.v conjugate logic.
    """
    re2 = ref_re
    im2 = -ref_im          # conjugate

    sub_re = int(re1) * int(re2) - int(im1) * int(im2)
    add_im = int(im1) * int(re2) + int(re1) * int(im2)

    re_out = (sub_re >> 14) + ((sub_re >> 13) & 1)
    im_out = (add_im >> 14) + ((add_im >> 13) & 1)

    re_out = _sign16(int(re_out) & 0xFFFF)
    im_out = _sign16(int(im_out) & 0xFFFF)
    return re_out, im_out


# ===========================================================================
# EXCEL DUMP HELPERS
# ===========================================================================

# Colour palette for stage header rows
_STAGE_COLOURS = [
    "1F4E79",   # 0_Inputs          — dark blue
    "375623",   # 1_FFT_Out         — dark green
    "7030A0",   # 2_MidBitRev_Out   — purple
    "C55A11",   # 3a_OFDM_Bins      — burnt orange
    "833C00",   # 3b_Radar_Bins     — dark orange
    "2E75B6",   # 4_ConjMul_Out     — mid blue
    "538135",   # 5_Shift_Out       — mid green
    "BF9000",   # 6_IFFT_Out        — gold
    "C00000",   # 7_FinalBitRev     — red
]

_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _make_header(ws, columns: list, colour_hex: str):
    """Write a bold white header row with a coloured background."""
    fill = PatternFill("solid", fgColor=colour_hex)
    font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    align = Alignment(horizontal="center", vertical="center")
    for col_idx, title in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font  = font
        cell.fill  = fill
        cell.alignment = align
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(title) + 2)
    ws.row_dimensions[1].height = 18


def _write_complex_sheet(ws, data: np.ndarray, colour_hex: str,
                          label_re="Real (int)", label_im="Imag (int)",
                          extra_cols: list = None, extra_data: list = None):
    """
    Write index + Re + Im columns (plus any extras) for a complex array.
    extra_cols : list of header strings for additional columns
    extra_data : list of lists (one per extra column, same length as data)
    """
    cols = ["Sample #", label_re, label_im]
    if extra_cols:
        cols += extra_cols
    _make_header(ws, cols, colour_hex)

    alt_fill  = PatternFill("solid", fgColor="EBF3FB")
    norm_font = Font(name="Arial", size=9)
    num_align = Alignment(horizontal="right")

    for i, val in enumerate(data):
        row = i + 2
        re  = int(np.real(val))
        im  = int(np.imag(val))
        fill = alt_fill if (i % 2 == 0) else None
        row_vals = [i, re, im]
        if extra_cols and extra_data:
            row_vals += [col[i] for col in extra_data]
        for col_idx, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.font   = norm_font
            cell.border = _BORDER
            cell.alignment = num_align
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"


def _write_split_sheet(ws, data_a: np.ndarray, data_b: np.ndarray,
                        colour_hex: str,
                        label_a="OFDM", label_b="Radar"):
    """
    Two side-by-side tables on one sheet: OFDM bins | Radar bins.
    """
    cols = [
        f"{label_a} Sample #", f"{label_a} Real", f"{label_a} Imag",
        "",   # spacer
        f"{label_b} Sample #", f"{label_b} Real", f"{label_b} Imag",
    ]
    _make_header(ws, cols, colour_hex)

    alt_fill  = PatternFill("solid", fgColor="EBF3FB")
    norm_font = Font(name="Arial", size=9)
    num_align = Alignment(horizontal="right")

    n = max(len(data_a), len(data_b))
    for i in range(n):
        row  = i + 2
        fill = alt_fill if (i % 2 == 0) else None
        row_vals = []
        if i < len(data_a):
            row_vals += [i, int(np.real(data_a[i])), int(np.imag(data_a[i]))]
        else:
            row_vals += ["", "", ""]
        row_vals.append("")    # spacer column
        if i < len(data_b):
            row_vals += [i, int(np.real(data_b[i])), int(np.imag(data_b[i]))]
        else:
            row_vals += ["", "", ""]
        for col_idx, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.font      = norm_font
            cell.border    = _BORDER
            cell.alignment = num_align
            if fill and col_idx != 4:
                cell.fill = fill

    ws.freeze_panes = "A2"


def _write_inputs_sheet(ws, rx_re, rx_im, ref_re, ref_im, colour_hex: str):
    """
    Sheet 0: rx_in (4096 rows) and ref_ram (2048 rows) side by side.
    Ref columns show blank after sample 2047.
    """
    cols = [
        "rx_in Sample #", "rx_in Real", "rx_in Imag",
        "",
        "ref_ram Addr", "ref_ram Real", "ref_ram Imag",
    ]
    _make_header(ws, cols, colour_hex)

    alt_fill  = PatternFill("solid", fgColor="EBF3FB")
    norm_font = Font(name="Arial", size=9)
    num_align = Alignment(horizontal="right")

    for i in range(N_FFT):
        row  = i + 2
        fill = alt_fill if (i % 2 == 0) else None
        row_vals = [i, int(rx_re[i]), int(rx_im[i]), ""]
        if i < N_HALF:
            row_vals += [i, int(ref_re[i]), int(ref_im[i])]
        else:
            row_vals += ["", "", ""]
        for col_idx, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.font      = norm_font
            cell.border    = _BORDER
            cell.alignment = num_align
            if fill and col_idx != 4:
                cell.fill = fill

    ws.freeze_panes = "A2"


def _dump_pipeline_xlsx(path: str, stages: dict):
    """
    Write all stage data to a formatted workbook.

    stages keys (in order):
        "inputs"       → (rx_re, rx_im, ref_re, ref_im)
        "fft_out"      → complex ndarray [4096]
        "mid_rev_out"  → complex ndarray [4096]
        "ofdm_bins"    → complex ndarray [2048]
        "radar_bins"   → complex ndarray [2048]
        "mul_out"      → complex ndarray [2048]  (before >>>3)
        "shift_out"    → complex ndarray [2048]  (after >>>3 = IFFT input)
        "ifft_out"     → complex ndarray [2048]
        "radar_final"  → complex ndarray [2048]
    """
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    c = _STAGE_COLOURS

    # ── Sheet 0: Inputs ───────────────────────────────────────────────────
    ws = wb.create_sheet("0_Inputs")
    rx_re, rx_im, ref_re, ref_im = stages["inputs"]
    _write_inputs_sheet(ws, rx_re, rx_im, ref_re, ref_im, c[0])

    # ── Sheet 1: FFT output ───────────────────────────────────────────────
    ws = wb.create_sheet("1_FFT_Out")
    _write_complex_sheet(ws, stages["fft_out"], c[1],
                         label_re="FFT Real (Q8.8)", label_im="FFT Imag (Q8.8)")

    # ── Sheet 2: Mid bit-reversal output ─────────────────────────────────
    ws = wb.create_sheet("2_MidBitRev_Out")
    _write_complex_sheet(ws, stages["mid_rev_out"], c[2],
                         label_re="BitRev Real", label_im="BitRev Imag")

    # ── Sheet 3: Demux — OFDM and Radar side by side ─────────────────────
    ws = wb.create_sheet("3_Demux_Out")
    _write_split_sheet(ws, stages["ofdm_bins"], stages["radar_bins"], c[3],
                       label_a="OFDM", label_b="Radar")

    # ── Sheet 4: Conjugate multiply output (before shift) ─────────────────
    ws = wb.create_sheet("4_ConjMul_Out")
    _write_complex_sheet(ws, stages["mul_out"], c[5],
                         label_re="Mul Real (raw)", label_im="Mul Imag (raw)")

    # ── Sheet 5: After >>>3 shift (IFFT input) ────────────────────────────
    ws = wb.create_sheet("5_Shift_Out")
    _write_complex_sheet(ws, stages["shift_out"], c[6],
                         label_re="Shifted Real (Q11.5)", label_im="Shifted Imag (Q11.5)")

    # ── Sheet 6: IFFT output ──────────────────────────────────────────────
    ws = wb.create_sheet("6_IFFT_Out")
    _write_complex_sheet(ws, stages["ifft_out"], c[7],
                         label_re="IFFT Real", label_im="IFFT Imag")

    # ── Sheet 7: Final bit-reversal (radar output) ────────────────────────
    ws = wb.create_sheet("7_FinalBitRev")
    _write_complex_sheet(ws, stages["radar_final"], c[8],
                         label_re="Radar Out Real", label_im="Radar Out Imag")

    wb.save(path)
    print(f"[DEBUG DUMP] Pipeline workbook saved  {path}")


# ===========================================================================
# MASTER PIPELINE FUNCTION
# ===========================================================================

def run_rx_pipeline(
    rx_re_array,
    rx_im_array,
    ref_re_array,
    ref_im_array,
    out_ofdm_re_file   = "rx_golden_ofdm_re.txt",
    out_ofdm_im_file   = "rx_golden_ofdm_im.txt",
    out_radar_re_file  = "rx_golden_radar_re.txt",
    out_radar_im_file  = "rx_golden_radar_im.txt",
    debug_xlsx         = "rx_pipeline_debug.xlsx",   # set None to disable
):
    print("==================================================")
    print("  STARTING RX TOP-LEVEL GOLDEN PIPELINE")
    print("==================================================")

    rx_re  = np.array(rx_re_array,  dtype=np.int64)
    rx_im  = np.array(rx_im_array,  dtype=np.int64)
    ref_re = np.array(ref_re_array, dtype=np.int64)
    ref_im = np.array(ref_im_array, dtype=np.int64)

    if len(rx_re) != N_FFT or len(rx_im) != N_FFT:
        raise ValueError(f"rx arrays must be length {N_FFT}, got {len(rx_re)}")
    if len(ref_re) != N_HALF or len(ref_im) != N_HALF:
        raise ValueError(f"ref arrays must be length {N_HALF}, got {len(ref_re)}")

    # ── Stage 1: FFT ──────────────────────────────────────────────────────
    print(">> Stage 1: Running bit-true 4096-pt FFT...")
    fft_out = radix2_dif_fft_fixed(rx_re + 1j * rx_im, WL=WL, FL=FL_FFT, is_ifft=False)
    print(f"   FFT first 4 bins: {fft_out[:4]}")

    # ── Stage 2: Mid-path bit-reversal (ADDR_W=12) ────────────────────────
    print(">> Stage 2: Mid-path bit-reversal (ADDR_W=12)...")
    mid_rev_out = _apply_bit_reversal(fft_out, addr_w=12)
    print(f"   mid_rev first 4 samples: {mid_rev_out[:4]}")

    # ── Stage 3: ISAC Demux ───────────────────────────────────────────────
    print(">> Stage 3: ISAC Demux split...")
    ofdm_bins  = mid_rev_out[:N_HALF]
    radar_bins = mid_rev_out[N_HALF:]
    print(f"   OFDM bin[0]  = {ofdm_bins[0]}")
    print(f"   Radar bin[0] = {radar_bins[0]}")

    # ── Stage 4: Conjugate multiply ───────────────────────────────────────
    print(">> Stage 4: Complex conjugate multiply...")
    mul_out   = np.zeros(N_HALF, dtype=np.complex128)
    shift_out = np.zeros(N_HALF, dtype=np.complex128)
    for k in range(N_HALF):
        re1 = int(np.real(radar_bins[k]))
        im1 = int(np.imag(radar_bins[k]))
        re_out, im_out = _multiply_rtl(re1, im1, int(ref_re[k]), int(ref_im[k]))
        mul_out[k] = complex(re_out, im_out)

        # ── Stage 5: >>>3 shift (captured separately for debug) ──────────
        shift_out[k] = complex(_asr(re_out, 3), _asr(im_out, 3))

    print(f"   Multiply product[0]: {mul_out[0]}")

    # ── Stage 6: IFFT ─────────────────────────────────────────────────────
    print(">> Stage 5: Running bit-true 2048-pt IFFT...")
    ifft_out = radix2_dif_ifft_fixed(shift_out, WL=WL, FL=FL_IFFT)
    print(f"   IFFT first 4 samples: {ifft_out[:4]}")

    # ── Stage 7: Final bit-reversal (ADDR_W=11) ───────────────────────────
    print(">> Stage 6: Final bit-reversal (ADDR_W=11)...")
    radar_final = _apply_bit_reversal(ifft_out, addr_w=11)
    print(f"   Radar output[0]: {radar_final[0]}")

    # ── Extract integer output arrays ─────────────────────────────────────
    ofdm_out_re  = np.real(ofdm_bins).astype(np.int64)
    ofdm_out_im  = np.imag(ofdm_bins).astype(np.int64)
    radar_out_re = np.real(radar_final).astype(np.int64)
    radar_out_im = np.imag(radar_final).astype(np.int64)

    # ── Export golden text files ──────────────────────────────────────────
    print(">> Exporting golden outputs for UVM scoreboard...")
    np.savetxt(out_ofdm_re_file,  ofdm_out_re,  fmt='%d')
    np.savetxt(out_ofdm_im_file,  ofdm_out_im,  fmt='%d')
    np.savetxt(out_radar_re_file, radar_out_re, fmt='%d')
    np.savetxt(out_radar_im_file, radar_out_im, fmt='%d')

    print(f"\n[SUCCESS] RX Pipeline Complete! Outputs saved to:")
    print(f"   - {out_ofdm_re_file}")
    print(f"   - {out_ofdm_im_file}")
    print(f"   - {out_radar_re_file}")
    print(f"   - {out_radar_im_file}")

    # ── Optional per-stage debug dump ─────────────────────────────────────
    if debug_xlsx is not None:
        _dump_pipeline_xlsx(debug_xlsx, {
            "inputs"      : (rx_re, rx_im, ref_re, ref_im),
            "fft_out"     : fft_out,
            "mid_rev_out" : mid_rev_out,
            "ofdm_bins"   : ofdm_bins,
            "radar_bins"  : radar_bins,
            "mul_out"     : mul_out,
            "shift_out"   : shift_out,
            "ifft_out"    : ifft_out,
            "radar_final" : radar_final,
        })

    return ofdm_out_re, ofdm_out_im, radar_out_re, radar_out_im


# ===========================================================================
# SELF-TEST
# ===========================================================================
if __name__ == "__main__":
    rng = np.random.default_rng(42)
    rx_re  = rng.integers(-32768, 32767, N_FFT,  dtype=np.int64)
    rx_im  = rng.integers(-32768, 32767, N_FFT,  dtype=np.int64)
    ref_re = rng.integers(-32768, 32767, N_HALF, dtype=np.int64)
    ref_im = rng.integers(-32768, 32767, N_HALF, dtype=np.int64)

    try:
        o_re, o_im, r_re, r_im = run_rx_pipeline(rx_re, rx_im, ref_re, ref_im)
        print(f"\nOFDM  output sample #0:    re={o_re[0]}  im={o_im[0]}")
        print(f"Radar output sample #0:    re={r_re[0]}  im={r_im[0]}")
        print(f"OFDM  output sample #2047: re={o_re[2047]}  im={o_im[2047]}")
        print(f"Radar output sample #2047: re={r_re[2047]}  im={r_im[2047]}")
    except Exception as e:
        print(f"\n[ERROR] Pipeline failed: {e}")
        raise
