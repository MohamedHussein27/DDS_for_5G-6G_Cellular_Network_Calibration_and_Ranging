"""
    Sponsor: Analog Devices, Inc. (ADI)
    Institution: Faculty of Engineering, Ain Shams University
    Project: DDS for 5G/6G Cellular Network Calibration and Ranging

    Module: system_top_golden_model.py

    Description:
        Master wrapper for the COMPLETE ISAC system golden model.
        Chains the full TX datapath directly into the full RX datapath
        with no channel in between (direct loopback connection).

        TX output is fed bit-for-bit as RX input:
            rx_in_re = tx_out_re
            rx_in_im = tx_out_im

    ┌─────────────────────────────────────────────────────────────────┐
    │                    FULL SYSTEM PIPELINE                          │
    │                                                                  │
    │  ┌─── TX PATH ─────────────────────────────────────────────┐   │
    │  │  DDS → FFT(4096) → MUX(OFDM + Radar>>7) → IFFT(4096)   │   │
    │  │         │                                                 │   │
    │  │   tx_spectrum[4096]  (kept for ref_ram build)            │   │
    │  └─────────────────────┬───────────────────────────────────┘   │
    │            tx_out[4096] (direct wire, no channel)               │
    │                         │                                       │
    │  ┌─── REF RAM BUILD ────┴──────────────────────────────────┐   │
    │  │  Inverse-permute TX spectrum through bit_rev_12 to align │   │
    │  │  with RX demux output order → ref_ram[2048]              │   │
    │  └─────────────────────┬───────────────────────────────────┘   │
    │                         │ rx_in = tx_out  +  ref_ram[2048]     │
    │                         ▼                                       │
    │  ┌─── RX PATH ─────────────────────────────────────────────┐   │
    │  │  FFT(4096) → BitRev(12) → Demux                          │   │
    │  │    → OFDM out [2048]   (communication signal)            │   │
    │  │    → ConjMul(ref_ram) → >>>3 → IFFT(2048) → BitRev(11)  │   │
    │  │    → Radar range profile [2048]                          │   │
    │  └──────────────────────────────────────────────────────────┘   │
    └─────────────────────────────────────────────────────────────────┘

    Reference RAM construction:
        The TX MUX places radar chirp (>>7 scaled) in bins [2429..4095]
        of the 4096-bin spectrum.  After RX FFT(4096) + mid-path
        bit-reversal (ADDR_W=12), the demux splits at sample 2048, so:

            radar_bins[r] corresponds to TX bin bit_rev_12(2048 + r)

        ref_ram[r] is filled with X_combined[bit_rev_12(2048 + r)],
        giving the correlator the exact transmitted radar amplitude at
        each frequency position — no additional scaling needed.

    Debug Excel dump (debug_xlsx != None):
        One workbook, 15 sheets — every TX and RX pipeline boundary
        plus a system summary sheet with scalar metrics.
"""

import numpy as np
from openpyxl                        import Workbook
from openpyxl.styles                 import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils                  import get_column_letter

from dds_new_model  import generate_dds_golden_model
from fft_fixed      import radix2_dif_fft_fixed
from ifft_fixed     import radix2_dif_ifft_fixed
from mux_new        import mux
from rx_golden import run_rx_pipeline


# ===========================================================================
# PARAMETERS
# ===========================================================================
WL      = 16
FL_FFT  = 8
FL_IFFT = 5
N_FFT   = 4096
N_HALF  = 2048
N_RADAR = 1667     # active radar bins placed by TX MUX (bins 2429..4095)


# ===========================================================================
# HELPERS
# ===========================================================================

def _sign16(v: int) -> int:
    v = int(v) & 0xFFFF
    return v - 0x10000 if (v & 0x8000) else v


def _asr(val: int, shift: int) -> int:
    """Arithmetic shift right — mirrors Verilog >>>."""
    return int(val) >> shift


def _bit_rev(k: int, width: int) -> int:
    """Reverse 'width' LSBs of k."""
    result = 0
    for _ in range(width):
        result = (result << 1) | (k & 1)
        k >>= 1
    return result


# ===========================================================================
# REFERENCE RAM CONSTRUCTION
# ===========================================================================

def build_ref_ram(tx_spectrum: np.ndarray) -> tuple:
    """
    Build ref_ram[2048] from the 4096-bin TX MUX output spectrum.

    Mapping derivation:
        RX: radar_bins[r] = mid_rev_out[2048 + r]
                          = fft_out[ bit_rev_12(2048 + r) ]
        Loopback: rx_fft == tx_spectrum  →  tx bin = bit_rev_12(2048 + r)

    ref_ram[r] = tx_spectrum[ bit_rev_12(2048 + r) ]
    Already carries the >>7 scaling applied by the TX MUX.
    The RX multiplier uses ref values raw (no additional shift).

    Parameters
    ----------
    tx_spectrum : complex ndarray [4096]  — X_combined from TX MUX output

    Returns
    -------
    ref_re, ref_im : int64 ndarray [2048]
    """
    ref_re = np.zeros(N_HALF, dtype=np.int64)
    ref_im = np.zeros(N_HALF, dtype=np.int64)

    for r in range(N_HALF):
        tx_bin    = _bit_rev(2048 + r, width=12)
        sample    = tx_spectrum[tx_bin]
        ref_re[r] = int(np.real(sample))
        ref_im[r] = int(np.imag(sample))

    non_zero = np.count_nonzero(ref_re) + np.count_nonzero(ref_im)
    print(f"   ref_ram built — non-zero entries: {non_zero}/{N_HALF}")
    return ref_re, ref_im


# ===========================================================================
# EXCEL DUMP
# ===========================================================================

_THIN   = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _make_header(ws, columns, colour_hex):
    fill  = PatternFill("solid", fgColor=colour_hex)
    font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    align = Alignment(horizontal="center", vertical="center")
    for ci, title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=title)
        cell.font = font; cell.fill = fill
        cell.alignment = align; cell.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = max(18, len(title) + 2)
    ws.row_dimensions[1].height = 18


def _write_complex_sheet(ws, data, colour_hex, label_re="Real (int)", label_im="Imag (int)"):
    _make_header(ws, ["Sample #", label_re, label_im], colour_hex)
    alt  = PatternFill("solid", fgColor="EBF3FB")
    font = Font(name="Arial", size=9)
    aln  = Alignment(horizontal="right")
    for i, val in enumerate(data):
        row  = i + 2
        fill = alt if (i % 2 == 0) else None
        for ci, v in enumerate([i, int(np.real(val)), int(np.imag(val))], 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = font; cell.border = _BORDER; cell.alignment = aln
            if fill: cell.fill = fill
    ws.freeze_panes = "A2"


def _write_summary_sheet(ws, metrics: list):
    _make_header(ws, ["Metric", "Value", "Notes"], "1F4E79")
    font = Font(name="Arial", size=10)
    for ri, (key, val, note) in enumerate(metrics, 2):
        for ci, v in enumerate([key, val, note], 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = font; cell.border = _BORDER
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EBF3FB")
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 45
    ws.freeze_panes = "A2"


def _dump_system_xlsx(path: str, tx_stages: dict, rx_stages: dict,
                       summary_metrics: list):
    """
    Single workbook:
        Sheet 00 : SYSTEM_Summary
        TX_01..06: TX stages
        RX_07..15: RX stages
    """
    wb = Workbook()
    wb.remove(wb.active)

    # ── Summary ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("00_SYSTEM_Summary")
    _write_summary_sheet(ws, summary_metrics)

    # ── TX stages ────────────────────────────────────────────────────────
    # DDS output (real only)
    ws = wb.create_sheet("TX_01_DDS_Out")
    _make_header(ws, ["Sample #", "Real (int)", "Imag (0)"], "1F4E79")
    font = Font(name="Arial",size=9); aln = Alignment(horizontal="right")
    alt  = PatternFill("solid", fgColor="EBF3FB")
    for i, v in enumerate(tx_stages["dds_out"]):
        row = i + 2; fill = alt if i % 2 == 0 else None
        for ci, cv in enumerate([i, int(v), 0], 1):
            cell = ws.cell(row=row, column=ci, value=cv)
            cell.font=font; cell.border=_BORDER; cell.alignment=aln
            if fill: cell.fill = fill
    ws.freeze_panes = "A2"

    ws = wb.create_sheet("TX_02_FFT_Out")
    _write_complex_sheet(ws, tx_stages["fft_out"], "003366",
                          "FFT Real (Q8.8)", "FFT Imag (Q8.8)")

    ws = wb.create_sheet("TX_03_MUX_Out")
    _write_complex_sheet(ws, tx_stages["mux_out"], "004080",
                          "MUX Real", "MUX Imag")

    ws = wb.create_sheet("TX_04_IFFT_Out")
    _write_complex_sheet(ws, tx_stages["tx_out"], "19598A",
                          "TX Out Real (int)", "TX Out Imag (int)")

    ws = wb.create_sheet("TX_05_RefRAM")
    ref_cplx = tx_stages["ref_re"] + 1j * tx_stages["ref_im"]
    _write_complex_sheet(ws, ref_cplx, "1B9CDB",
                          "Ref_RAM Real", "Ref_RAM Imag")

    # ── Direct wire — TX out == RX in (no sheet needed, just annotated) ──
    # RX stages ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("RX_06_FFT_Out")
    _write_complex_sheet(ws, rx_stages["fft_out"], "375623",
                          "FFT Real (Q8.8)", "FFT Imag (Q8.8)")

    ws = wb.create_sheet("RX_07_MidBitRev")
    _write_complex_sheet(ws, rx_stages["mid_rev_out"], "537D3A",
                          "BitRev Real", "BitRev Imag")

    # Demux — OFDM and Radar side by side
    ws = wb.create_sheet("RX_08_Demux_Out")
    _make_header(ws,
                 ["OFDM #","OFDM Real","OFDM Imag","",
                  "Radar #","Radar Real","Radar Imag"], "699B4A")
    font=Font(name="Arial",size=9); aln=Alignment(horizontal="right")
    alt=PatternFill("solid",fgColor="EBF3FB")
    ofdm_b=rx_stages["ofdm_bins"]; radar_b=rx_stages["radar_bins"]
    for i in range(N_HALF):
        row=i+2; fill=alt if i%2==0 else None
        row_vals=[i,int(np.real(ofdm_b[i])),int(np.imag(ofdm_b[i])),"",
                  i,int(np.real(radar_b[i])),int(np.imag(radar_b[i]))]
        for ci,cv in enumerate(row_vals,1):
            cell=ws.cell(row=row,column=ci,value=cv)
            cell.font=font; cell.border=_BORDER; cell.alignment=aln
            if fill and ci!=4: cell.fill=fill
    ws.freeze_panes="A2"

    ws = wb.create_sheet("RX_09_ConjMul_Out")
    _write_complex_sheet(ws, rx_stages["mul_out"], "2E75B6",
                          "Mul Real (raw)", "Mul Imag (raw)")

    ws = wb.create_sheet("RX_10_Shift_Out")
    _write_complex_sheet(ws, rx_stages["shift_out"], "C55A11",
                          "Shifted Real (Q11.5)", "Shifted Imag (Q11.5)")

    ws = wb.create_sheet("RX_11_IFFT_Out")
    _write_complex_sheet(ws, rx_stages["ifft_out"], "7030A0",
                          "IFFT Real", "IFFT Imag")

    ws = wb.create_sheet("RX_12_Radar_Final")
    _write_complex_sheet(ws, rx_stages["radar_final"], "C00000",
                          "Radar Out Real", "Radar Out Imag")

    ws = wb.create_sheet("RX_13_OFDM_Final")
    _write_complex_sheet(ws, rx_stages["ofdm_bins"], "BF9000",
                          "OFDM Out Real", "OFDM Out Imag")

    wb.save(path)
    print(f"[DEBUG DUMP] System workbook saved → {path}")


# ===========================================================================
# MASTER SYSTEM PIPELINE FUNCTION
# ===========================================================================

def run_system_top_pipeline(
    # ── TX parameters ──────────────────────────────────────────────────
    FTW_start     = 0,
    FTW_step      = 426666,
    N_cycles      = 4096,
    Fs            = 491.52e6,
    ofdm_re_array = None,   # list/ndarray [2048] — OFDM ROM data (real)
    ofdm_im_array = None,   # list/ndarray [2048] — OFDM ROM data (imag)

    # ── Output files ───────────────────────────────────────────────────
    out_tx_re_file    = "sys_golden_tx_re.txt",
    out_tx_im_file    = "sys_golden_tx_im.txt",
    out_ofdm_re_file  = "sys_golden_ofdm_re.txt",
    out_ofdm_im_file  = "sys_golden_ofdm_im.txt",
    out_radar_re_file = "sys_golden_radar_re.txt",
    out_radar_im_file = "sys_golden_radar_im.txt",

    # ── Debug ──────────────────────────────────────────────────────────
    debug_xlsx    = "system_pipeline_debug.xlsx",   # None to disable
):
    """
    Run the full ISAC system golden model with TX wired directly to RX.

        rx_in_re = tx_out_re
        rx_in_im = tx_out_im   (bit-for-bit, no channel)

    Returns
    -------
    tx_out_re, tx_out_im       : int64 ndarray [4096]
    ofdm_out_re, ofdm_out_im   : int64 ndarray [2048]
    radar_out_re, radar_out_im : int64 ndarray [2048]
    """
    print("==================================================")
    print("  STARTING FULL ISAC SYSTEM GOLDEN PIPELINE")
    print("  TX → (direct wire) → RX")
    print("==================================================")

    # ─────────────────────────────────────────────────────────────────────
    # TX STAGE 1: DDS
    # ─────────────────────────────────────────────────────────────────────
    print("\n>>> [TX] Stage 1: Bit-True DDS...")
    chirp_time = generate_dds_golden_model(FTW_start, FTW_step, N_cycles)
    x_re = np.real(chirp_time).astype(np.int64).flatten()
    x_im = np.imag(chirp_time).astype(np.int64).flatten()

    # Enforce exactly 4096 samples
    L = min(len(x_re), len(x_im))
    if L < N_FFT:
        x_re = np.pad(x_re, (0, N_FFT - L), 'constant')
        x_im = np.pad(x_im, (0, N_FFT - L), 'constant')
    else:
        x_re, x_im = x_re[:N_FFT], x_im[:N_FFT]

    chirp_input = x_re + 1j * x_im
    print(f"   DDS last 4 samples: {chirp_time[-4:]}")

    # ─────────────────────────────────────────────────────────────────────
    # TX STAGE 2: FFT
    # ─────────────────────────────────────────────────────────────────────
    print(">>> [TX] Stage 2: FFT(4096)...")
    chirp_freq = radix2_dif_fft_fixed(chirp_input, WL=WL, FL=FL_FFT, is_ifft=False)
    stream_neg_scaled = chirp_freq[:N_RADAR]   # first 1667 bins into MUX
    print(f"   chirp_freq[0] = {chirp_freq[0]}")

    # ─────────────────────────────────────────────────────────────────────
    # TX STAGE 3: OFDM array
    # ─────────────────────────────────────────────────────────────────────
    print(">>> [TX] Stage 3: Fetching OFDM data...")
    if ofdm_re_array is None or ofdm_im_array is None:
        print("   [WARNING] No OFDM arrays provided — using zeros (radar-only mode).")
        ofdm_re_int = np.zeros(N_HALF, dtype=np.int64)
        ofdm_im_int = np.zeros(N_HALF, dtype=np.int64)
    else:
        ofdm_re_int = np.array(ofdm_re_array, dtype=np.int64)
        ofdm_im_int = np.array(ofdm_im_array, dtype=np.int64)
        # Mirror one-cycle pipeline delay in the ROM read logic
        ofdm_re_int[1:] = ofdm_re_int[:-1]
        ofdm_im_int[1:] = ofdm_im_int[:-1]

    stream_pos_int = ofdm_re_int + 1j * ofdm_im_int
    print(f"   OFDM[0] = re={ofdm_re_int[0]}  im={ofdm_im_int[0]}")

    # ─────────────────────────────────────────────────────────────────────
    # TX STAGE 4: MUX
    # ─────────────────────────────────────────────────────────────────────
    print(">>> [TX] Stage 4: 3-State MUX...")
    X_combined, _ = mux(stream_pos_int, stream_neg_scaled, N_cycles)
    print(f"   MUX[0]    = {X_combined[0]}")
    print(f"   MUX[2048] = {X_combined[2048]}  (first zero-pad bin)")
    print(f"   MUX[2429] = {X_combined[2429]}  (first radar bin, >>7 applied)")

    # ─────────────────────────────────────────────────────────────────────
    # TX STAGE 5: IFFT → tx_out
    # ─────────────────────────────────────────────────────────────────────
    print(">>> [TX] Stage 5: IFFT(4096)...")
    tx_time   = radix2_dif_ifft_fixed(X_combined, WL=WL, FL=FL_IFFT)
    tx_out_re = np.real(tx_time).astype(np.int64)
    tx_out_im = np.imag(tx_time).astype(np.int64)
    print(f"   tx_out[0] = re={tx_out_re[0]}  im={tx_out_im[0]}")
    print(f"   tx_out[4095] = re={tx_out_re[-1]}  im={tx_out_im[-1]}")

    # ─────────────────────────────────────────────────────────────────────
    # REFERENCE RAM — built from TX MUX spectrum (X_combined)
    # ─────────────────────────────────────────────────────────────────────
    print(">>> [SYS] Building Reference RAM from TX spectrum...")
    ref_re, ref_im = build_ref_ram(X_combined)
    print(f"   ref_ram[0]    = re={ref_re[0]}  im={ref_im[0]}")
    print(f"   ref_ram[2047] = re={ref_re[-1]} im={ref_im[-1]}")

    # ─────────────────────────────────────────────────────────────────────
    # DIRECT WIRE: rx_in = tx_out  (no channel)
    # ─────────────────────────────────────────────────────────────────────
    print(">>> [SYS] Direct wire: rx_in = tx_out (no channel)")
    rx_in_re = tx_out_re   # same object — no copy needed
    rx_in_im = tx_out_im

    # ─────────────────────────────────────────────────────────────────────
    # RX PIPELINE
    # ─────────────────────────────────────────────────────────────────────
    print(">>> [RX]  Running full RX pipeline...")
    ofdm_out_re, ofdm_out_im, radar_out_re, radar_out_im = run_rx_pipeline(
        rx_re_array   = rx_in_re,
        rx_im_array   = rx_in_im,
        ref_re_array  = ref_re,
        ref_im_array  = ref_im,
        out_ofdm_re_file  = out_ofdm_re_file,
        out_ofdm_im_file  = out_ofdm_im_file,
        out_radar_re_file = out_radar_re_file,
        out_radar_im_file = out_radar_im_file,
        debug_xlsx    = None,   # consolidated into system workbook below
    )

    # ─────────────────────────────────────────────────────────────────────
    # TX OUTPUT EXPORT
    # ─────────────────────────────────────────────────────────────────────
    np.savetxt(out_tx_re_file, tx_out_re, fmt='%d')
    np.savetxt(out_tx_im_file, tx_out_im, fmt='%d')

    # ─────────────────────────────────────────────────────────────────────
    # SCALAR METRICS
    # ─────────────────────────────────────────────────────────────────────
    tx_power   = float(np.mean(tx_out_re.astype(float)**2 + tx_out_im.astype(float)**2))
    radar_mag  = np.sqrt(radar_out_re.astype(float)**2 + radar_out_im.astype(float)**2)
    peak_bin   = int(np.argmax(radar_mag))
    peak_amp   = float(radar_mag[peak_bin])
    p2avg_dB   = float(10 * np.log10(peak_amp**2 / (np.mean(radar_mag**2) + 1e-12)))
    ofdm_nz    = int(np.count_nonzero(ofdm_out_re) + np.count_nonzero(ofdm_out_im))

    print("\n──────────────────────────────────────────────────")
    print(f"  TX signal power            : {tx_power:.2f} LSB²")
    print(f"  Radar peak bin             : {peak_bin}")
    print(f"  Radar peak amplitude       : {peak_amp:.2f} LSB")
    print(f"  Radar peak-to-avg (dB)     : {p2avg_dB:.2f} dB")
    print(f"  OFDM non-zero output bins  : {ofdm_nz}/{N_HALF*2}")
    print("──────────────────────────────────────────────────\n")

    # ─────────────────────────────────────────────────────────────────────
    # DEBUG XLSX — reconstruct per-stage RX arrays for dump
    # ─────────────────────────────────────────────────────────────────────
    if debug_xlsx is not None:

        def _apply_br(s, w):
            n = len(s); o = np.zeros(n, dtype=np.complex128)
            for k in range(n): o[k] = s[_bit_rev(k, w)]
            return o

        def _mul_rtl(re1, im1, rr, ri):
            im2 = -ri
            sub = int(re1)*int(rr) - int(im1)*int(im2)
            add = int(im1)*int(rr) + int(re1)*int(im2)
            ro  = (sub >> 14) + ((sub >> 13) & 1)
            io_ = (add >> 14) + ((add >> 13) & 1)
            return _sign16(int(ro) & 0xFFFF), _sign16(int(io_) & 0xFFFF)

        rx_cplx     = rx_in_re + 1j * rx_in_im
        fft_out_rx  = radix2_dif_fft_fixed(rx_cplx, WL=WL, FL=FL_FFT, is_ifft=False)
        mid_rev     = _apply_br(fft_out_rx, 12)
        ofdm_bins   = mid_rev[:N_HALF]
        radar_bins  = mid_rev[N_HALF:]
        mul_out     = np.zeros(N_HALF, dtype=np.complex128)
        shift_out   = np.zeros(N_HALF, dtype=np.complex128)
        for k in range(N_HALF):
            ro, io = _mul_rtl(int(np.real(radar_bins[k])), int(np.imag(radar_bins[k])),
                               int(ref_re[k]), int(ref_im[k]))
            mul_out[k]   = complex(ro, io)
            shift_out[k] = complex(_asr(ro, 3), _asr(io, 3))
        ifft_out_rx = radix2_dif_ifft_fixed(shift_out, WL=WL, FL=FL_IFFT)
        radar_final = _apply_br(ifft_out_rx, 11)

        summary = [
            ("Connection",              "TX direct wire to RX",  "No channel — rx_in = tx_out"),
            ("FTW_start",               FTW_start,               ""),
            ("FTW_step",                FTW_step,                ""),
            ("N_cycles",                N_cycles,                ""),
            ("Fs (MHz)",                f"{Fs/1e6:.3f}",         ""),
            ("TX Signal Power (LSB²)",  f"{tx_power:.2f}",       ""),
            ("Radar Peak Bin",          peak_bin,                 ""),
            ("Radar Peak Amplitude",    f"{peak_amp:.2f}",        "LSB"),
            ("Radar Peak-to-Avg (dB)",  f"{p2avg_dB:.2f}",       ""),
            ("OFDM Non-Zero Out Bins",  ofdm_nz,                  f"out of {N_HALF*2}"),
            ("OFDM ROM supplied",       str(ofdm_re_array is not None), ""),
        ]

        _dump_system_xlsx(
            debug_xlsx,
            tx_stages = {
                "dds_out" : chirp_time,
                "fft_out" : chirp_freq,
                "mux_out" : X_combined,
                "tx_out"  : tx_time,
                "ref_re"  : ref_re,
                "ref_im"  : ref_im,
            },
            rx_stages = {
                "fft_out"    : fft_out_rx,
                "mid_rev_out": mid_rev,
                "ofdm_bins"  : ofdm_bins,
                "radar_bins" : radar_bins,
                "mul_out"    : mul_out,
                "shift_out"  : shift_out,
                "ifft_out"   : ifft_out_rx,
                "radar_final": radar_final,
            },
            summary_metrics = summary,
        )

    print("[SUCCESS] Full system pipeline complete!")
    print(f"   TX     → {out_tx_re_file}, {out_tx_im_file}")
    print(f"   OFDM   → {out_ofdm_re_file}, {out_ofdm_im_file}")
    print(f"   Radar  → {out_radar_re_file}, {out_radar_im_file}")

    return tx_out_re, tx_out_im, ofdm_out_re, ofdm_out_im, radar_out_re, radar_out_im


# ===========================================================================
# SELF-TEST
# ===========================================================================
if __name__ == "__main__":
    rng = np.random.default_rng(0)

    ofdm_re = rng.integers(-16384, 16383, N_HALF, dtype=np.int64).tolist()
    ofdm_im = rng.integers(-16384, 16383, N_HALF, dtype=np.int64).tolist()

    try:
        tx_re, tx_im, o_re, o_im, r_re, r_im = run_system_top_pipeline(
            ofdm_re_array = ofdm_re,
            ofdm_im_array = ofdm_im,
            debug_xlsx    = "system_pipeline_debug.xlsx",
        )
        print(f"\nTX     [0]:    re={tx_re[0]}    im={tx_im[0]}")
        print(f"OFDM   [0]:    re={o_re[0]}     im={o_im[0]}")
        print(f"Radar  [0]:    re={r_re[0]}     im={r_im[0]}")
        print(f"OFDM   [2047]: re={o_re[2047]}  im={o_im[2047]}")
        print(f"Radar  [2047]: re={r_re[2047]}  im={r_im[2047]}")
    except Exception as e:
        print(f"[ERROR] {e}"); raise
